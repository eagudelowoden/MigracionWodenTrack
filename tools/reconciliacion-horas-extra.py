#!/usr/bin/env python3
"""
Reconciliación de horas extra: impacto de mover la jornada nocturna de 21:00 → 19:00.

Cuantifica, por colaborador, cuántas horas trabajadas caen en la franja 19:00–21:00
(que es la ÚNICA diferencia entre el límite viejo de 21:00 y el nuevo de 19:00 de la
Ley 2466 de 2025) y estima el impacto en recargo usando los factores de ley.

Estas horas, al pasar la noche de 21:00 a 19:00, cambian de categoría:
  - Dentro de turno, día ordinario:      (sin recargo) → RN     (+0.35 por hora)
  - Dentro de turno, festivo/domingo:    RDDF (0.75)   → RNDF   (+0.35 por hora)
  - Fuera de turno (extra), día ordinario: HEDO (1.25)  → HENO  (+0.50 por hora)
  - Fuera de turno (extra), festivo:       HEFD (2.00)  → HEFN  (+0.50 por hora)

El "delta factor" multiplicado por (valor hora ordinaria del colaborador) da el
dinero adicional a pagar. Como el Excel no trae salarios, se reporta en horas y en
"delta-factor-hora" (equivalente a fracciones de hora ordinaria); pásale --valor-hora
para estimar pesos con un valor único, o deja el cálculo de pesos para nómina.

Uso:
  python tools/reconciliacion-horas-extra.py "ruta/Consolidado.xlsx"
  python tools/reconciliacion-horas-extra.py "ruta/Consolidado.xlsx" --valor-hora 6189   # 2026 aprox SMLMV/240
  python tools/reconciliacion-horas-extra.py "ruta/Consolidado.xlsx" --csv salida.csv
"""
import sys
import argparse
from datetime import date, timedelta

try:
    import openpyxl
except ImportError:
    print("Falta openpyxl. Instala con: pip install openpyxl")
    sys.exit(1)

# ── Festivos Colombia (mismo algoritmo que el backend) ───────────────────────
def easter(year):
    a = year % 19; b = year // 100; c = year % 100
    d = b // 4; e = b % 4; f = (b + 8) // 25; g = (b - f + 1) // 3
    h = (19*a + b - d - g + 15) % 30; i = c // 4; k = c % 4
    l = (32 + 2*e + 2*i - h - k) % 7; m = (a + 11*h + 22*l) // 451
    month = (h + l - 7*m + 114) // 31; day = ((h + l - 7*m + 114) % 31) + 1
    return date(year, month, day)

def next_monday(d):
    return d + timedelta(days=(8 - d.isoweekday()) % 7)

def festivos_colombia(year):
    f = set()
    for mes, dia in [(1,1),(5,1),(7,20),(8,7),(12,8),(12,25)]:
        f.add(date(year, mes, dia))
    for mes, dia in [(1,6),(3,19),(6,29),(8,15),(10,12),(11,1),(11,11)]:
        f.add(next_monday(date(year, mes, dia)))
    e = easter(year)
    # relativos a Pascua (Emiliani los traslada a lunes salvo los de Semana Santa)
    f.add(e + timedelta(days=-3))   # Jueves Santo
    f.add(e + timedelta(days=-2))   # Viernes Santo
    f.add(next_monday(e + timedelta(days=43)))  # Ascensión
    f.add(next_monday(e + timedelta(days=64)))  # Corpus Christi
    f.add(next_monday(e + timedelta(days=71)))  # Sagrado Corazón
    return f

# ── Solapamiento de minutos ──────────────────────────────────────────────────
def overlap(s, e, a, b):
    return max(0, min(e, b) - max(s, a))

def to_min(v):
    if v is None:
        return None
    if hasattr(v, "hour"):
        return v.hour * 60 + v.minute
    s = str(v).strip().upper()
    ampm = 0
    if s.endswith("AM") or s.endswith("PM"):
        if s.endswith("PM"):
            ampm = 12
        s = s[:-2].strip()
    try:
        p = s.split(":")
        h = int(p[0]); mi = int(p[1]) if len(p) > 1 else 0
        if ampm == 12 and h != 12:
            h += 12
        if ampm == 0 and h == 12 and str(v).strip().upper().endswith("AM"):
            h = 0
        return h * 60 + mi
    except Exception:
        return None

# Minutos en la franja 19:00–21:00 dentro de [ing, sal] (soporta cruce de medianoche)
def mins_19_21(ing, sal):
    if sal <= ing:
        sal += 1440
    total = overlap(ing, sal, 1140, 1260)          # 19:00–21:00 mismo día
    if sal > 1440:                                  # porción después de medianoche
        total += overlap(0, sal - 1440, 1140, 1260)
    return total

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="Ruta del Consolidado .xlsx")
    ap.add_argument("--valor-hora", type=float, default=None,
                    help="Valor de la hora ordinaria para estimar pesos (opcional)")
    ap.add_argument("--csv", default=None, help="Exportar resultado a CSV")
    ap.add_argument("--hoja", default=None, help="Nombre de la hoja (por defecto la primera)")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    ws = wb[args.hoja] if args.hoja else wb.worksheets[0]

    # Columnas (1-based) según el consolidado:
    # A1 ced  B2 nombre  C3 fecha  D4 turnoIni  E5 turnoFin  F6 ingreso  G7 salida
    festivos_cache = {}
    def es_festivo(d):
        if d.year not in festivos_cache:
            festivos_cache[d.year] = festivos_colombia(d.year)
        return d.isoweekday() == 7 or d in festivos_cache[d.year]

    # delta-factor por hora según categoría (al pasar de 21:00 a 19:00)
    DELTA = {
        "ord_dentro":  0.35,   # → RN
        "fest_dentro": 0.35,   # RDDF → RNDF
        "ord_extra":   0.50,   # HEDO → HENO
        "fest_extra":  0.50,   # HEFD → HEFN
    }

    por_colab = {}   # ced -> {nombre, horas, delta_factor, dias}
    total_filas = 0
    total_min_1921 = 0

    for r in range(3, ws.max_row + 1):
        ced = ws.cell(r, 1).value
        nombre = ws.cell(r, 2).value
        fecha = ws.cell(r, 3).value
        ing = to_min(ws.cell(r, 6).value)
        sal = to_min(ws.cell(r, 7).value)
        if ced is None or ing is None or sal is None:
            continue
        total_filas += 1

        m = mins_19_21(ing, sal)
        if m <= 0:
            continue
        total_min_1921 += m
        horas = m / 60.0

        # ¿festivo?
        d = None
        if hasattr(fecha, "year"):
            d = date(fecha.year, fecha.month, fecha.day)
        festivo = es_festivo(d) if d else False

        # ¿dentro de turno? si hay turno (D/E) y la franja cae dentro del turno → dentro
        t_ini = to_min(ws.cell(r, 4).value)
        t_fin = to_min(ws.cell(r, 5).value)
        dentro = False
        if t_ini is not None and t_fin is not None:
            tf = t_fin if t_fin > t_ini else t_fin + 1440
            # ¿la franja 19-21 está dentro del turno?
            franja_ini = max(ing, 1140)
            franja_fin = min(sal if sal > ing else sal + 1440, 1260)
            if overlap(franja_ini, franja_fin, t_ini, tf) > 0:
                dentro = True

        key = ("fest" if festivo else "ord") + ("_dentro" if dentro else "_extra")
        delta = DELTA[key] * horas

        rec = por_colab.setdefault(str(ced), {"nombre": nombre or "", "horas": 0.0,
                                              "delta_factor": 0.0, "dias": 0})
        rec["horas"] += horas
        rec["delta_factor"] += delta
        rec["dias"] += 1

    # ── Salida ───────────────────────────────────────────────────────────────
    filas = sorted(por_colab.items(), key=lambda x: x[1]["delta_factor"], reverse=True)

    print("=" * 92)
    print("RECONCILIACIÓN: impacto de mover la jornada nocturna de 21:00 a 19:00")
    print("=" * 92)
    print(f"Filas procesadas: {total_filas}")
    print(f"Total horas en franja 19:00–21:00: {total_min_1921/60:.2f} h")
    print(f"Colaboradores afectados: {len(filas)}")
    if args.valor_hora:
        total_pesos = sum(v["delta_factor"] for _, v in filas) * args.valor_hora
        print(f"Impacto total estimado: ${total_pesos:,.0f}  (valor hora ${args.valor_hora:,.0f})")
    print("-" * 92)
    hdr = f"{'CÉDULA':>14} {'NOMBRE':<34} {'DÍAS':>5} {'H 19-21':>9} {'Δ-FACTOR':>9}"
    if args.valor_hora:
        hdr += f" {'~PESOS':>12}"
    print(hdr)
    print("-" * 92)
    for ced, v in filas:
        line = f"{ced:>14} {str(v['nombre'])[:34]:<34} {v['dias']:>5} {v['horas']:>9.2f} {v['delta_factor']:>9.2f}"
        if args.valor_hora:
            line += f" {v['delta_factor']*args.valor_hora:>12,.0f}"
        print(line)

    if args.csv:
        import csv
        with open(args.csv, "w", newline="", encoding="utf-8-sig") as fh:
            w = csv.writer(fh)
            cols = ["cedula", "nombre", "dias", "horas_19_21", "delta_factor"]
            if args.valor_hora:
                cols.append("pesos_estimados")
            w.writerow(cols)
            for ced, v in filas:
                row = [ced, v["nombre"], v["dias"], round(v["horas"], 2), round(v["delta_factor"], 2)]
                if args.valor_hora:
                    row.append(round(v["delta_factor"] * args.valor_hora))
                w.writerow(row)
        print(f"\nCSV exportado: {args.csv}")

    print("\nNOTA: 'Δ-FACTOR' = fracciones de hora ordinaria adicionales a pagar por el")
    print("cambio de categoría (no incluye el valor base de la hora, solo el recargo extra).")
    print("Multiplícalo por el valor-hora real de cada colaborador para el monto exacto.")

if __name__ == "__main__":
    main()
